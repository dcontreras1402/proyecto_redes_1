import csv
from openpyxl import load_workbook

def procesar_dataset_productos():
    vendedores = [17, 21, 26, 29, 33, 39, 43, 47, 51, 55, 59, 63, 69, 73, 78, 81, 85, 89, 93, 97, 102, 105, 109, 113, 117]
    idx_vendedor = 0

    ruta_entrada = "/vagrant/proyecto_redes_1/Electonic Shop Dataset.xlsx"
    ruta_salida = "/vagrant/proyecto_redes_1/productos_limpios.csv"

    productos_unicos = {}

    wb = load_workbook(filename=ruta_entrada, read_only=True)
    hoja = wb.active

    iterable_hoja = hoja.iter_rows(values_only=True)
    cabecera = next(iterable_hoja)
    
    idx_nombre = cabecera.index("product_name")
    idx_precio = cabecera.index("product_price")

    for fila in iterable_hoja:
        if not fila or fila[idx_nombre] is None:
            continue
        nombre = str(fila[idx_nombre]).strip()
        precio = str(fila[idx_precio]).strip()

        if nombre and nombre not in productos_unicos:
            productos_unicos[nombre] = {
                'precio': precio,
                'descripcion': f"Excelente {nombre} disponible en tienda electronica."
            }

    lista_productos = list(productos_unicos.items())
    total_unicos = len(lista_productos)

    with open(ruta_salida, mode='w', encoding='utf-8', newline='') as f_out:
        escritor = csv.writer(f_out)
        escritor.writerow(["id_vendedor", "nombre", "descripcion", "precio", "cantidad", "activo", "aprobado"])

        contador = 0
        while contador < 10000:
            nombre_base, datos = lista_productos[contador % total_unicos]

            vendedor_id = vendedores[idx_vendedor]
            idx_vendedor = (idx_vendedor + 1) % len(vendedores)

            sufijo = f" Lote {str(contador // total_unicos + 1)}" if contador >= total_unicos else ""
            nombre_final = f"{nombre_base}{sufijo}"
            descripcion_final = f"{datos['descripcion']} ID Registro: {contador + 1}."

            cantidad = str((contador % 150) + 10)
            precio_final = datos['precio']

            escritor.writerow([vendedor_id, nombre_final, descripcion_final, precio_final, cantidad, "1", "1"])
            contador += 1

if __name__ == "__main__":
    procesar_dataset_productos()
